import sys
import re
import openpyxl

#Imputs
txt_path = sys.argv[1]
xlsx_path = sys.argv[2]
tipo = sys.argv[3]
'''
Tipo:
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/Banco do Brasil janeiro 2022  cc 51734-8.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/Banco do Brasil janeiro 2022  cc 51734-8.xlsx" "Banco do Brasil"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/Bradesco dez.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/Bradesco dez.xlsx" "Bradesco"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/BS2 887946-0.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/BS2 887946-0.xlsx" "BS2"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/c6.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/c6.xlsx" "C6"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/caixa292-5 - julho-22 (1).txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/caixa292-5 - julho-22 (1).xlsx" "Caixa"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/202402 - CITIBANK_90015799 Extrato de OperaÃ§Ãµes (1)_ORC.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/202402 - CITIBANK_90015799 Extrato de OperaÃ§Ãµes (1)_ORC.xlsx" "City"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/ItauAbril2023.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/ItauAbril2023.xlsx" "Itau"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/Original Fev.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/Original Fev.xlsx" "Original"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/santander1.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/santander_ORC.xlsx" "Santander"
python TxtToXlsx.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/travelex_ORC.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/travelex_ORC.xlsx" "Travelex"
'''

# Ler o texto do arquivo 
with open(txt_path, 'r', encoding='utf-8') as file:
    texto = file.read()

    # Variáveis da função
    matches = None
    colunas = None
    matchesCabecalho = None
    matchesCorpo = None
    ordemColunas = None

match tipo:
    case "Banco do Brasil":
        # Padrão regex
        padrao = r"(\d{2}/\d{2}/\d{4}) (\d{4}) (\d{5}) ([\w\d/. ]+) ([\d.]+) ([ R$\d,.-]+,\d{2}\b) (\w)\n([\w\d./ -]+)"
        
        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Agência Origem', 'lote', 'Histórico', 'Documento', 'Valor R$', 'Operador', 'Descrição']

    case "Bradesco":
        # Padrão regex
        padrao = r"(\d{2}/\d{2}/\d{4})*\n*([\w\d./ -]*\n[\w\d./ -]*)\n(\d+) ([-.\d,]+,\d{2}\b-*) ([-.\d,]+,\d{2}\b-*)"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Lançamento', 'Documento', 'Valor', 'Saldo']
        
    case "BS2":
        # Padrão regex
        padrao = r"(\d{2}/\d{2}/\d{4}) ([\w\d./ ]+-*[\w\d./ ]+) ([R$ \d,.-]+,\d{2}\b)"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Descrição', 'Valor']
        
    case "C6":
        # Padrão regex
        padrao = r"(\d{2}/\d{2}/\d{4}) ([\w ]+) ([-.\d,]+,\d{2}\b)"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Descrição', 'Valor']
        
    case "Caixa":
        # Padrão regex
        padrao = r"(\d{2}/\d{2}/\d{4}) (\d{6}) ([\w\d./ -]*) ([-.\d,]+,\d{2}\b-*) (\w) ([-.\d,]+,\d{2}\b-*) (\w)"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Nº Documento', 'Histórico', 'Valor', 'Operador Valor', 'Saldo', 'Operador Saldo']
        
    case "City":
        # Padrão regex
        padrao = r"(\d{16})\s*[|]\s*([\w\s./\-]*)\s*[|]\s*(\d{2}/\d{2}/\d{4})\s*[|]\s*(\d{2}/\d{2}/\d{4})\s*[|]\s*(\d{2}/\d{2}/\d{4})\s*[|]\s*([\w\s./\-]*)\s+([\d,\d{3}.-]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))\s*([-.\d,\d{2}]*(?=))"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Número da Operação', 'Título', 'Data de Início', 'Data de Vencimento', 'Data de Liquidação', 'Indexador', '(%) do Indexador', 'Taxa Original (a.a)', 'Valor Inicial da Aplicação (R$)', 'Valor Base da Aplicação Corrigido (R$)', 'Rendimento Bruto do Título (R$)', 'IOF (R$)', 'IRRF (R$)', 'Rendimento Líquido do Título (R$)', 'Valor Base de Aplicação Liquido (R$)', '% Resgate Antecipado']
        
    case "Itau":
        # Padrão regex
        padraoCabecalho = r"(D = débito a compensar )+(\d{2}/\d{2})([ \w\d./-]*(?=)) ([-.\d,\d{2}]*(?=))|(G = aplicação programada )+([ \w\d./-]*(?=)) ([-.\d,\d{2}]*(?=))|(P = poupança automática )+([ \w\d./-]*(?=)) ([-.\d,\d{2}]*(?=))"
        padraoCorpo = r"^(?!.*SALDO APLIC AUT MAIS)(\d{2}/\d{2})*([a-zA-Z]+[\w\d./ -]+) +([-.\d,]+,\d{2}\b-*)\s"

        # Encontrar o índice do trecho "Para demais siglas, consulte as Notas"
        indice_split = texto.find("Para demais siglas, consulte as Notas")
        parte1 = texto[:indice_split]
        parte2 = texto[indice_split:]

        # Encontrar o índice do trecho "totalizador de aplicações automáticas entrada R$ saída R$"
        indice_split2 = parte2.find("totalizador de aplicações automáticas entrada R$ saída R$")
        parte2Usavel = parte2[:indice_split2]

        # Procurando por todas as correspondências no texto
        matchesCabecalho = re.findall(padraoCabecalho, parte1)
        matchesCorpo = re.findall(padraoCorpo, parte2Usavel, re.MULTILINE)

        # Define padrão de colunas
        colunas = ['Data', 'Descrição', 'Valor']
        
    case "Original":
        # Padrão regex
        padrao = r"([\w\d./ -]*)\n(\d{2}/\d{2}/\d{4}) (\w+) ([+-])* [R$]+ ([-.\d,]+,\d{2}\b)\n([\w\d./ -]*)"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Lançamento', 'tipo', 'Operador', 'Valor', 'Descrição']

    case "Santander":
        # Padrão regex
        padrao = r"(\d{2}/\d{2}/\d{4})\s*([\w\s\d./-]*(?=))\s*(\d{6})\s*([\d,\d{2}.-]*(?=))"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Histórico', 'Nº Documento', 'Valor']

    case "Travelex":
        # Padrão regex
        padrao = r"\d{2}/\d{2}/\d{4}\n(\d{2}/\d{2}/\d{4})\n\d{2}:\d{2}\n([\w\sç]*\n(?!.*LIQUIDACAO)\w*)\n([\w\s\d\n./:|-]*(?=))\b(?:Sim|Não)\b ([-.\d,\d{2}]*(?=)) ([-.\d,\d{2}]*(?=))"

        # Procurando por todas as correspondências no texto
        matches = re.findall(padrao, texto)

        # Define padrão de colunas
        colunas = ['Data', 'Tipo', 'Detalhes', 'Valor', 'Saldo']
                                  
    case _:
        raise ValueError("Tipo inválido!")


if tipo == "Itau":
    # Se houver correspondências, escrever os dados em um arquivo XLSX
    if matchesCabecalho and matchesCorpo:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(colunas)
        for match in matchesCabecalho:
            ws.append([match[1], match[2], match[3]])
            ws.append(['', match[5], match[6]])
            ws.append(['', match[8], match[9]])

        # Remover linhas vazias
        rows_to_delete = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if all(not cell.value for cell in row):
                rows_to_delete.add(row[0].row)
        for row_index in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_index, amount=1)

        for match in matchesCorpo:
            ws.append([match[0], match[1], match[2]])

        # Percorrer as células na coluna C e converter os valores para números com duas casas decimais
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                if cell.value:
                    # Remover pontos e vírgulas
                    cleaned_value = cell.value.replace('.', '').replace(',', '')
                    # Verificar se o último caractere é "-" e converter para negativo se necessário
                    if cleaned_value.endswith('-'):
                        cleaned_value = '-' + cleaned_value[:-1]
                    # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                    float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                    cell.value = float_value

        # Preencher células vazias na coluna "Data" com valores das células acima até encontrar outra célula não vazia
        last_value = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value:
                    last_value = cell.value
                else:
                    cell.value = last_value

        wb.save(xlsx_path)
        print(f"Arquivo XLSX criado com sucesso em: {xlsx_path}")
    else:
        print("Nenhuma correspondência encontrada.")

else:
    # Se houver correspondências, escrever os dados em um arquivo XLSX
    if matches:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(colunas)
        if tipo == "Original":
            for match in matches:
                ws.append(list([match[1], match[0], match[2], match[3], match[4], match[5]]))
        else:
            for match in matches:
                ws.append(list(match))

        if tipo == 'Bradesco':
            # Preencher células vazias na coluna "Data" com valores das células acima até encontrar outra célula não vazia
            last_value = None
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value:
                        last_value = cell.value
                    else:
                        cell.value = last_value

        # Percorrer as células nas colunas de valor e converter os valores para números com duas casas decimais
        if tipo == "Banco do Brasil":
            for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
                cell6 = row[0]  # Primeira célula (coluna 6)
                cell7 = row[1]  # Segunda célula (coluna 7)
                if cell7.value == "D":
                    if cell6.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell6.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        # Fazer o número negativo
                        cell6.value = -float_value
                else:
                    if cell6.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell6.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        # Fazer o número negativo
                        cell6.value = float_value
        if tipo == "Bradesco":
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
        if tipo == "BS2":
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('R$', '').replace('.', '').replace(',', '')
                        # Verificar se o último caractere é "-" e converter para negativo se necessário
                        if cleaned_value.startswith('-') or cleaned_value.startswith('--') :
                            cleaned_value = '-' + cleaned_value[3:]
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
        if tipo == "C6":
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
        if tipo == "Caixa":
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
                cell4 = row[0]  # Primeira célula (coluna 6)
                cell5 = row[1]  # Segunda célula (coluna 7)
                if cell5.value == "D":
                    if cell4.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell4.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        # Fazer o número negativo
                        cell4.value = -float_value
                else:
                    if cell4.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell4.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        # Fazer o número positivo
                        cell4.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
                cell6 = row[0]  # Primeira célula (coluna 6)
                cell7 = row[1]  # Segunda célula (coluna 7)
                if cell7.value == "D":
                    if cell6.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell6.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        # Fazer o número negativo
                        cell6.value = -float_value
                else:
                    if cell6.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell6.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        # Fazer o número positivo
                        cell6.value = float_value
        if tipo == "City":
            for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=13, max_col=13):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=14, max_col=14):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
            for row in ws.iter_rows(min_row=2, min_col=15, max_col=15):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
        if tipo == "Original":
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
        if tipo == "Santander":
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value
        if tipo == "Travelex": 
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value         
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    if cell.value:
                        # Remover pontos e vírgulas
                        cleaned_value = cell.value.replace('.', '').replace(',', '')
                        # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                        float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                        cell.value = float_value  
        
        wb.save(xlsx_path)
        print(f"Arquivo XLSX criado com sucesso em: {xlsx_path}")
    else:
        print("Nenhuma correspondência encontrada.")