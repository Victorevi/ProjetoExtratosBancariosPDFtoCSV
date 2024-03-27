import sys
import re
import openpyxl

txt_path = sys.argv[1]

# Ler o texto do arquivo 
with open(txt_path, 'r', encoding='utf-8') as file:
    texto = file.read()
    
# Padrão regex
padraoCabecalho = r"(D = débito a compensar )+(\d{2}/\d{2})([ \w\d./-]*(?=)) ([-.\d,\d{2}]*(?=))|(G = aplicação programada )+([ \w\d./-]*(?=)) ([-.\d,\d{2}]*(?=))|(P = poupança automática )+([ \w\d./-]*(?=)) ([-.\d,\d{2}]*(?=))"
padraoCorpo = r"(\d{2}/\d{2})*([\w\d./ -]*) ([-.\d,]+,\d{2}\b-*)\s"


# Encontrar o índice do trecho "Para demais siglas, consulte as Notas"
indice_split = texto.find("Para demais siglas, consulte as Notas")
parte1 = texto[:indice_split]
parte2 = texto[indice_split:]
#print(texto)


# Procurando por todas as correspondências no texto
matchesCabecalho = re.findall(padraoCabecalho, parte1)
matchesCorpo = re.findall(padraoCorpo, parte2)
'''for match in matchesCorpo:
        print(match[0])'''


# Se houver correspondências, escrever os dados em um arquivo XLSX
if matchesCabecalho and matchesCorpo:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Data', 'Descrição', 'Valor'])
    for match in matchesCabecalho:
        ws.append([match[1], match[2], match[3]])
        ws.append(['', match[5], match[6]])
        ws.append(['', match[8], match[9]])

    # Remover linhas vazias
    rows_to_delete = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if all(not cell.value for cell in row):
            rows_to_delete.add(row[0].row)
    for row_index in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_index, amount=1)

    for match in matchesCorpo:
        ws.append([match[0], match[1], match[2]])

    # Percorrer as células na coluna C e converter os valores para números com duas casas decimais
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value:
                # Remover pontos e vírgulas
                cleaned_value = cell.value.replace('.', '').replace(',', '')
                # Verificar se o último caractere é "-" e converter para negativo se necessário
                if cleaned_value.endswith('-'):
                    cleaned_value = '-' + cleaned_value[:-1]
                # Converter para float com duas casas decimais e ajustar as duas últimas casas decimais como parte decimal
                float_value = round(float(cleaned_value[:-2] + '.' + cleaned_value[-2:]), 2)
                cell.value = float_value

    # Preencher células vazias na coluna "Data" com valores das células acima até encontrar outra célula não vazia
    last_value = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                last_value = cell.value
            else:
                cell.value = last_value

    xlsx_path = sys.argv[2]
    wb.save(xlsx_path)
    print(f"Arquivo XLSX criado com sucesso em: {xlsx_path}")
else:
    print("Nenhuma correspondência encontrada.")

#python converte_itau.py "C:/Users/vbarbosa/Downloads/docs bancarios/Script/ItauAbril2023.txt" "C:/Users/vbarbosa/Downloads/docs bancarios/Script/ItauAbril2023.xlsx"