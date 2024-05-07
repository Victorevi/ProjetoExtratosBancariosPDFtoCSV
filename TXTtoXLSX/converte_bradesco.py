import sys
import re
import openpyxl

txt_path = sys.argv[1]

# Ler o texto do arquivo 
with open(txt_path, 'r', encoding='utf-8') as file:
    texto = file.read()
    
# Padrão regex
padrao = r"(\d{2}/\d{2}/\d{4})*\n*([\w\d./ -]*\n[\w\d./ -]*)\n(\d+) ([-.\d,]+,\d{2}\b-*) ([-.\d,]+,\d{2}\b-*)"

# Procurando por todas as correspondências no texto
matches = re.findall(padrao, texto)

# Se houver correspondências, escrever os dados em um arquivo XLSX
if matches:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Data', 'Lançamento', 'Documento', 'Valor', 'Saldo'])
    for match in matches:
        ws.append(list(match))
    
    # Preencher células vazias na coluna "Data" com valores das células acima até encontrar outra célula não vazia
    last_value = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                last_value = cell.value
            else:
                cell.value = last_value

    xlsx_path = sys.argv[2]
    wb.save(xlsx_path)
    print(f"Arquivo XLSX criado com sucesso em: {xlsx_path}")
else:
    print("Nenhuma correspondência encontrada.")